import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from io import BytesIO

def export_grades_template(exam, students):
    """
    Génère un fichier Excel servant de modèle pour l'importation des notes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Importation des notes"

    # En-têtes
    headers = [
        "Matricule", 
        "Nom de l'étudiant", 
        "Note (Max: {})".format(exam.max_score), 
        "Absent (O/N)", 
        "Remarques"
    ]
    
    # Style pour les en-têtes
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Données des étudiants
    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student.student_id)
        ws.cell(row=row, column=2, value=student.user.get_full_name())
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value="N")
        ws.cell(row=row, column=5, value="")

    ws.column_dimensions['B'].width = 35

    # Instructions
    ws_instr = wb.create_sheet(title="Instructions")
    ws_instr.cell(row=1, column=1, value="Instructions pour l'importation").font = Font(bold=True, size=14)
    ws_instr.cell(row=3, column=1, value="1. Ne pas modifier les colonnes Matricule et Nom.")
    ws_instr.cell(row=4, column=1, value="2. Saisir les notes dans la colonne C.")
    ws_instr.cell(row=5, column=1, value="3. Pour les absents, mettre 'O' dans la colonne D.")
    ws_instr.cell(row=6, column=1, value="4. La note maximale autorisée pour cet examen est: {}".format(exam.max_score))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_current_grades(exam, grades_data):
    """
    Génère un fichier Excel avec les notes actuelles.
    grades_data: list of dicts with student and score info
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notes actuelles"

    headers = ["Matricule", "Nom", "Note", "Absent", "Remarques", "Noté par", "Date"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    for row, grade in enumerate(grades_data, 2):
        ws.cell(row=row, column=1, value=grade['student_matricule'])
        ws.cell(row=row, column=2, value=grade['student_name'])
        ws.cell(row=row, column=3, value=grade['score'])
        ws.cell(row=row, column=4, value="Oui" if grade['is_absent'] else "Non")
        ws.cell(row=row, column=5, value=grade.get('remarks', ''))
        ws.cell(row=row, column=6, value=grade['graded_by_name'])
        ws.cell(row=row, column=7, value=grade['graded_at'])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
